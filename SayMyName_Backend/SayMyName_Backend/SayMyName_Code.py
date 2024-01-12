import openpyxl as excel;             #For reading the excel file
from g2p_en import G2p;
from flask_cors import CORS
from flask import Flask, request, jsonify
from flask_restful import Api
import json
import re

app = Flask(__name__)
api = Api(app)
CORS(app)

def getRequiredPhonetics(user_name, student_id):
        #Reading the excel file 
        students = r'E:\AIGS\1003 - Machine Learning\SayMyName_Backend\SayMyName_Backend\Student_details.xlsx'
        # students = r'Student_details.xlsx'
        wb_obj = excel.load_workbook(students)
        sheet = wb_obj['Sheet1']
        row = sheet.max_row         #Stores the max value of the row
        #g2p initialization
        g2p = G2p()
        #Function for extracting the phonetics
        def phonetics(i):
                cell_name = sheet.cell(row = i, column = 2)
                # using g2p for generating phonetics
                phonetic_name = g2p(cell_name.value.lower())
                # appending the string from the list generated above of phonetics
                word = ''
                for x in phonetic_name:
                        word = word + x
                # taking out number from the string
                        true_phonetices = ''.join("-" if c.isdigit() else c for c in word)
                        if true_phonetices[len(true_phonetices)-1] == "-":
                                true_phonetices = true_phonetices[:-1] + ""
                        cell_phn = sheet.cell(row= i, column=3)
                        if cell_name.value.lower() == user_name.lower():
                                cell_phn.value = true_phonetices.lower()
                                wb_obj.save(students)
                return true_phonetices.lower()

        for i in range(1, row + 1):  
                cell_id = sheet.cell(row = i, column = 1)
                if cell_id.value == int(student_id):
                        cell_name = sheet.cell(row = i, column = 2).value
                        if (cell_name.lower() == user_name.lower()): 
                                cell_phn = sheet.cell(row = i, column = 3).value
                                if (cell_phn != None):
                                        requestObj = {
                                                        "phoneticsSpelling": cell_phn,
                                                        "status": "success",
                                                        "message": "success" 
                                                }
                                        return requestObj
                                else:
                                        true_phonetices = phonetics(i)
                                        requestObj = {
                                                        "phoneticsSpelling": true_phonetices.lower(),
                                                        "status": "success",
                                                        "message": "success" 
                                                }
                                        return requestObj
                        else:
                                requestObj = {
                                                        "phoneticsSpelling": None,
                                                        "status": "failed",
                                                        "message": "The username entered does not match the existing studentID. Please enter the correct name." 
                                                }
                                return requestObj             
        #This block of code executes if the input name does not exist in the Excel
        else:
                #This block of code executes once we reached the last row and we didn't find the name
                if i == row:
                #This loop executes and checks for the 2nd column which is the student ID
                        for j in range(1, row+1):
                                cell_name = sheet.cell(row = j+1, column = 2)
                        #This block of code executes if the entered student ID matches with any other student's id
                                if (cell_name.value == user_name.lower()):
                                        requestObj = {
                                                "phoneticsSpelling": None,
                                                "status": "failed",
                                                "message": f"Mismatch between studentID and student name. The ID {cell_id.value} already exists for other student in the database."
                                                }
                                        return requestObj
                        #This block of code will store the Name and Student ID if it is not already present in the Excel and will speak the Name
                                else:
                                        j = j + 1
                                        if j == row:
                                                sheet.append([int(student_id), user_name])
                                                wb_obj.save(students)
                                                i = i + 1
                                                j = j + 1
                                                row = sheet.max_row 
                                                true_phonetices = phonetics(i)
                                                requestObj = {
                                                        "phoneticsSpelling": true_phonetices.lower(),
                                                        "status": "success",
                                                        "message": "success" 
                                                }
                                                return requestObj


def updateUserFeedback(user_name, student_id, feedback):
        students = r'E:\AIGS\1003 - Machine Learning\SayMyName_Backend\SayMyName_Backend\Student_details.xlsx'
        # students = r'Student_details.xlsx'
        wb_obj = excel.load_workbook(students)
        sheet = wb_obj['Sheet1']
        row = sheet.max_row 
        for i in range(1, row + 1):  
                cell_id = sheet.cell(row = i, column = 1)
                if cell_id.value == int(student_id): # student id
                        cell_name = sheet.cell(row = i, column = 2).value
                        if (cell_name.lower() == user_name.lower()): #student name
                                cell_name = sheet.cell(row = i, column = 4)
                                cell_name.value = feedback
                                wb_obj.save(students)
                                requestObj = {
                                                        "status": "success",
                                                        "message": "success" 
                                                }
                                return requestObj
                                
                        else:
                                requestObj = {
                                                        "status": "failed",
                                                        "message": "Something went wrong" 
                                                }
                                return requestObj
                                           
        
        


@app.route('/getPhonetics', methods=['POST'])
def getPhonetics():
       record = json.loads(request.data)
       if (record["studentId"].isdigit() and bool(re.match("^[A-Za-z]*$",record["studentName"]))):
                reply = getRequiredPhonetics(record["studentName"], record["studentId"])
                return jsonify(reply) 
       else:
                reqObj ={
                         "phoneticsSpelling": None,
                         "status": "failed",
                         "message": "Incorrect StudentID or Username"
                        }
                return jsonify(reqObj) 
       
@app.route('/userFeedback', methods=['POST'])
def getUserFeedback():
        record_data = json.loads(request.data)
        reply = updateUserFeedback(record_data["studentName"], record_data["studentId"], record_data['feedback'])
        return jsonify(reply) 

      

if __name__ == '__main__':
     app.run(port=5002)